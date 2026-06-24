"""CSV storage and exception logging for comment collection."""

from __future__ import annotations

import csv
import html
import zipfile
from dataclasses import asdict
from pathlib import Path
from typing import Iterable, Mapping, Sequence, Set

from .models import CommentRecord


COMMENT_FIELDNAMES = [
    "video_id",
    "comment_id",
    "root_comment_id",
    "parent_comment_id",
    "reply_to_comment_id",
    "reply_to_user_name",
    "depth",
    "comment_path",
    "comment_user_name",
    "comment_user_id_hash",
    "comment_user_uid_hash",
    "comment_time",
    "comment_ip_location",
    "comment_like_count",
    "comment_text",
    "cleaned_comment_text",
    "text_length",
    "has_emoji",
    "has_url",
    "has_mention",
    "crawl_time",
]


class CSVCommentStore:
    """Append-only CSV store that supports checkpoint-style de-duplication."""

    def __init__(self, output_path: Path | str) -> None:
        """Create a CSV store and prepare the parent output directory."""

        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def load_existing_comment_ids(self) -> Set[str]:
        """Read already-saved comment IDs so interrupted runs can resume safely."""

        if not self.output_path.exists():
            return set()

        with self.output_path.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            return {
                row["comment_id"]
                for row in reader
                if row.get("comment_id")
            }

    def append_records(self, records: Iterable[CommentRecord]) -> int:
        """Append records to CSV immediately and return the number written."""

        rows = [asdict(record) for record in records]
        if not rows:
            return 0

        should_write_header = not self.output_path.exists() or self.output_path.stat().st_size == 0
        with self.output_path.open("a", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=COMMENT_FIELDNAMES)
            if should_write_header:
                writer.writeheader()
            writer.writerows(rows)
            file.flush()
        return len(rows)


class CSVFailureLogger:
    """Append-only CSV logger for failed comment requests."""

    fieldnames = [
        "video_url",
        "video_id",
        "comment_id",
        "depth",
        "page",
        "error",
        "crawl_time",
    ]

    def __init__(self, log_path: Path | str) -> None:
        """Create an exception logger and prepare its parent directory."""

        self.log_path = Path(log_path)
        self.log_path.parent.mkdir(parents=True, exist_ok=True)

    def log_failure(
        self,
        *,
        video_url: str,
        video_id: str,
        comment_id: str,
        depth: int,
        page: int,
        error: str,
        crawl_time: str,
    ) -> None:
        """Append one failed request with context for later inspection."""

        should_write_header = not self.log_path.exists() or self.log_path.stat().st_size == 0
        with self.log_path.open("a", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=self.fieldnames)
            if should_write_header:
                writer.writeheader()
            writer.writerow(
                {
                    "video_url": video_url,
                    "video_id": video_id,
                    "comment_id": comment_id,
                    "depth": depth,
                    "page": page,
                    "error": error,
                    "crawl_time": crawl_time,
                }
            )
            file.flush()


def read_csv_rows(path: Path | str) -> list[dict[str, str]]:
    """Read CSV rows using utf-8-sig encoding."""

    csv_path = Path(path)
    if not csv_path.exists():
        return []
    with csv_path.open("r", newline="", encoding="utf-8-sig") as file:
        return list(csv.DictReader(file))


def write_csv_rows(
    path: Path | str,
    rows: Sequence[Mapping[str, object]],
    *,
    fieldnames: Sequence[str] = COMMENT_FIELDNAMES,
) -> None:
    """Rewrite CSV rows using utf-8-sig encoding."""

    csv_path = Path(path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def export_rows_to_xlsx(
    path: Path | str,
    rows: Sequence[Mapping[str, object]],
    *,
    fieldnames: Sequence[str] = COMMENT_FIELDNAMES,
    metadata: Mapping[str, object] | None = None,
) -> None:
    """Export rows to a simple XLSX workbook using only the standard library."""

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    metadata = metadata or {}
    sheet_rows = [list(fieldnames)] + [
        [row.get(field, "") for field in fieldnames]
        for row in rows
    ]
    metadata_rows = [["key", "value"]] + [[key, value] for key, value in metadata.items()]
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _content_types_xml())
        archive.writestr("_rels/.rels", _root_rels_xml())
        archive.writestr("xl/workbook.xml", _workbook_xml())
        archive.writestr("xl/_rels/workbook.xml.rels", _workbook_rels_xml())
        archive.writestr("xl/worksheets/sheet1.xml", _worksheet_xml(sheet_rows))
        archive.writestr("xl/worksheets/sheet2.xml", _worksheet_xml(metadata_rows))


def _worksheet_xml(rows: Sequence[Sequence[object]]) -> str:
    """Build minimal worksheet XML with inline strings."""

    row_xml = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for column_index, value in enumerate(row, start=1):
            ref = f"{_column_name(column_index)}{row_index}"
            cells.append(
                f'<c r="{ref}" t="inlineStr"><is><t>{html.escape(str(value))}</t></is></c>'
            )
        row_xml.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        '</worksheet>'
    )


def _column_name(index: int) -> str:
    """Return an Excel column name for a 1-based index."""

    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _content_types_xml() -> str:
    """Return XLSX content type metadata."""

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '</Types>'
    )


def _root_rels_xml() -> str:
    """Return root relationship XML."""

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>'
    )


def _workbook_xml() -> str:
    """Return workbook XML with comments and metadata sheets."""

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets>'
        '<sheet name="comments" sheetId="1" r:id="rId1"/>'
        '<sheet name="metadata" sheetId="2" r:id="rId2"/>'
        '</sheets>'
        '</workbook>'
    )


def _workbook_rels_xml() -> str:
    """Return workbook relationship XML."""

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>'
        '</Relationships>'
    )


class XLSXCommentStore:
    """Append-only Excel store that uses the same export fields as CSV."""

    def __init__(self, output_path: Path | str, sheet_name: str = "comments") -> None:
        """Create an XLSX store; requires the optional openpyxl dependency."""

        self.output_path = Path(output_path)
        self.sheet_name = sheet_name
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def load_existing_comment_ids(self) -> Set[str]:
        """Read already-saved comment IDs from the workbook for resume support."""

        if not self.output_path.exists():
            return set()

        openpyxl = self._load_openpyxl()
        workbook = openpyxl.load_workbook(self.output_path)
        if self.sheet_name not in workbook.sheetnames:
            return set()

        sheet = workbook[self.sheet_name]
        headers = [cell.value for cell in sheet[1]]
        try:
            comment_id_col = headers.index("comment_id") + 1
        except ValueError:
            return set()

        ids: Set[str] = set()
        for row in sheet.iter_rows(min_row=2, min_col=comment_id_col, max_col=comment_id_col):
            value = row[0].value
            if value:
                ids.add(str(value))
        return ids

    def append_records(self, records: Iterable[CommentRecord]) -> int:
        """Append comment rows to an XLSX workbook and return the number written."""

        rows = [asdict(record) for record in records]
        if not rows:
            return 0

        openpyxl = self._load_openpyxl()
        if self.output_path.exists():
            workbook = openpyxl.load_workbook(self.output_path)
            sheet = workbook[self.sheet_name] if self.sheet_name in workbook.sheetnames else workbook.create_sheet(self.sheet_name)
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = self.sheet_name
            sheet.append(COMMENT_FIELDNAMES)

        first_row = [cell.value for cell in sheet[1]]
        if first_row != COMMENT_FIELDNAMES:
            sheet.append(COMMENT_FIELDNAMES)

        for row in rows:
            sheet.append([row[field] for field in COMMENT_FIELDNAMES])
        workbook.save(self.output_path)
        return len(rows)

    @staticmethod
    def _load_openpyxl():
        """Import openpyxl lazily so CSV-only use has no Excel dependency."""

        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("Install the optional 'excel' dependency to export XLSX files.") from exc
        return openpyxl
